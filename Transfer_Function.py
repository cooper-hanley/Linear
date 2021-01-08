# -*- coding: utf-8 -*-
"""
Created on Thu Nov 18 11:13:43 2020

@author: cooper.hanley
"""
###########################################################################
import matplotlib.pyplot as plt
import xlrd
import numpy as np
import Read_Sensor
import openpyxl

######
wb = openpyxl.load_workbook(Read_Sensor.ave_filename)
ws = wb.worksheets[0]
xval = []
yval = []
stdval = []
for cell in ws['A']:
    xval.append(cell.value)
for cell in ws['B']:
    yval.append(cell.value)
for cell in ws['C']:
    stdval.append(cell.value)

# Create linear regression from averages
x = []
y = []
std = []
for item in range(1, len(xval)):
    x.append(float(xval[item]))
    y.append(float(yval[item]))
    std.append(float(stdval[item]))


fit = np.polyfit(x,y,1)
fit_fn = np.poly1d(fit)
mxb = str(np.poly1d(fit))
equation = 'y = ' + mxb.strip()
# Write Slope and Y-int to spreadsheet
ws['E1'].value = 'Slope'
ws['E2'].value = 'Y-int'
ws['F1'].value = fit[0]
ws['F2'].value = fit[1]
wb.save(Read_Sensor.ave_filename)
# Can set your own fonts. See title section below.
title_font = {'fontname': 'Arial', 'size': '20', 'color': 'black', 'weight': 'normal',
              'verticalalignment': 'bottom'}

# 1) Plot Backdrop Color:
fig, ax = plt.subplots(facecolor='gainsboro', figsize=(8, 5))
# 2) Plot Area Color:
ax.set_facecolor('snow')
# 3) Title:
ax.set_title('Sensor Output vs. xValues', **title_font)
# 4) X-axis Label:
ax.set_xlabel('xValues', fontsize=16, color='black')
# 5) Y-axis Label:
ax.set_ylabel('Sensor Output', fontsize=16, color='black')
# 6) 1st Data Series with Error Bars:
ax.errorbar(x,y, yerr = std, fmt = 'o', markersize = 2)
#ax.plot(x,y,marker = 'x', markersize = 5, linestyle = 'None')
# 7) 1st Data Series Trend Line:
ax.plot(x,fit_fn(x), '--k', linewidth = .5)
# 8) Graph Customizing:
ax.tick_params(labelcolor='black')
plt.grid(True)
# Adding equation to plot
x_loc = min(x)
y_loc = max(y)*0.75
ax.text(x_loc, y_loc,equation)
# 9) Legend
plt.legend(['Linear Regression', 'Test Data'], loc='upper left')
fig.tight_layout()
plt.show()
# 10) Save the figure and png
# plt.savefig('Comparison.png', facecolor=fig.get_facecolor(), edgecolor='none', dpi=240)








